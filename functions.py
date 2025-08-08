####IMPORT MODULES
from bs4 import BeautifulSoup
from datetime import datetime
import requests
import random
import openpyxl
import shutil
import os

####IMPORT LOCAL
#from header_list import headers

headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"}

###FUNCTIONS

def check_write_permission(path):
    if os.path.exists(path):
        return os.path.isfile(path) and os.access(path, os.W_OK)
    else:
        parent_dir = os.path.dirname(path) or '.'
        return os.access(parent_dir, os.W_OK)

def back_up(src):
    print("back up init")
    if check_write_permission(src):
        dir_name = os.path.dirname(src)
        base_name = os.path.basename(src)
        name, ext = os.path.splitext(base_name)
        date = datetime.now().strftime("%d_%m_%Y-%H:%M")
        dst = os.path.join(dir_name, f"{name}_backup_{date}{ext}")
        try:
            shutil.copy(src, dst)
            print(f"Successfully made a back up of '{src}'.")
        except Exception as e:
            print(f"Failed to copy file: {e}")
    else:
        print("No write permission for the destination.")


def make_url(name):
    url = f"https://www.finanzen.net/aktien/{name}-aktie"
    return url


def get_price(url):
    try:
        #response = requests.get(url, headers=random.choice(headers), timeout=10)
        response = requests.get(url, headers=headers, timeout=20)
        response.raise_for_status()  
        
    except requests.exceptions.HTTPError as e:
        if e.response.status_code == 404:  
            print(f"404: Skipping {url}")
            return None  
        else:
            raise  
            
    except (requests.exceptions.RequestException, 
            requests.exceptions.Timeout,
            requests.exceptions.SSLError) as e:
        print(f"Network error ({type(e).__name__}): Skipping {url}")
        return None
        
    try:
        soup = BeautifulSoup(response.text, 'html.parser')
        span_tag = soup.find('span', class_='snapshot__value')
        number = span_tag.get_text(strip=True)
        return number
        
    except AttributeError: 
        print(f"Missing price element: Skipping {url}")
        return None

def handle_file(path):

    ##get the excel-file as a workbook
    wb = openpyxl.load_workbook(path)
    print("workbook opened")
    ##we expect the list of names to operate on to be on the first sheet, column A
    sheet = wb.worksheets[0]

    #if it's the first time running, create a StockPriceUpdate-sheet
    if "StockPriceUpdate" not in wb.sheetnames:
        update_sheet = wb.create_sheet("StockPriceUpdate")
        first_time = True
        print("made a new StockPriceUpdate Sheet")
    #when it exists, we refer to it as update_sheet
    update_sheet = wb["StockPriceUpdate"]
    print("StockPriceUpdate Sheet already exists")
    #StockPriceUpdate will have the names in column A and the updated prices in column B. All old data will be at C onwards.
    update_sheet.insert_cols(idx=1)
    update_sheet.insert_cols(idx=2)
    print("inserted columns")
    ##update the header
    update_sheet["A1"] = "names"
    update_sheet["B1"] = f'{datetime.now().strftime("Data %d/%m/%Y %H:%M")}'
    print("updated header")
    

    for cell_value in sheet.iter_cols(min_col=1, max_col=1, min_row=2, values_only=True):
        for row_idx, stock_name in enumerate(cell_value, start=2):
            if stock_name is not None and stock_name != "NaN":
                url = make_url(stock_name)
                print(url)
                price = get_price(url)
                update_sheet[f"A{row_idx}"] = stock_name
                update_sheet[f"B{row_idx}"] = price
    
    wb.save(path)
    



    


